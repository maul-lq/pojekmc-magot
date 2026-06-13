from __future__ import annotations

import json
import math
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from backend.config import Settings, settings
from backend.model import Model


class PayloadError(ValueError):
    pass


def _number(value: Any, name: str, minimum: float, maximum: float) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise PayloadError(f"{name} harus berupa angka.")
    number = float(value)
    if not math.isfinite(number) or not minimum <= number <= maximum:
        raise PayloadError(f"{name} berada di luar rentang yang diizinkan.")
    return number


def validate_payload(payload: bytes | str | dict[str, Any]) -> dict[str, Any]:
    if isinstance(payload, bytes):
        if len(payload) > 1024:
            raise PayloadError("Payload MQTT melebihi 1.024 byte.")
        payload = payload.decode("utf-8")
    if isinstance(payload, str):
        if len(payload.encode("utf-8")) > 1024:
            raise PayloadError("Payload MQTT melebihi 1.024 byte.")
        try:
            payload = json.loads(payload)
        except json.JSONDecodeError as exc:
            raise PayloadError("Payload MQTT bukan JSON yang valid.") from exc
    if not isinstance(payload, dict):
        raise PayloadError("Payload MQTT harus berupa objek JSON.")

    required = {"temperature", "humidity", "gas", "buzzer"}
    if not required.issubset(payload):
        missing = ", ".join(sorted(required - set(payload)))
        raise PayloadError(f"Payload MQTT tidak lengkap: {missing}.")

    temperature = _number(payload["temperature"], "temperature", -40, 80)
    humidity = _number(payload["humidity"], "humidity", 0, 100)
    gas_float = _number(payload["gas"], "gas", 0, 4095)
    if not gas_float.is_integer():
        raise PayloadError("gas harus berupa bilangan bulat.")
    gas = int(gas_float)

    buzzer = payload["buzzer"]
    if not isinstance(buzzer, str) or buzzer.upper() not in {"ON", "OFF"}:
        raise PayloadError("buzzer harus bernilai ON atau OFF.")
    buzzer = buzzer.upper()

    has_problem = temperature == 0 and humidity == 0
    temperature_abnormal = False if has_problem else not (29 < temperature < 39)
    gas_abnormal = gas > 2000
    expected_buzzer = "ON" if temperature_abnormal or gas_abnormal else "OFF"

    return {
        "temperature": temperature,
        "humidity": humidity,
        "gas": gas,
        "buzzer": buzzer,
        "temperature_abnormal": temperature_abnormal,
        "gas_abnormal": gas_abnormal,
        "has_problem": has_problem,
        "buzzer_inconsistent": False if has_problem else buzzer != expected_buzzer,
        "received_at": datetime.now(timezone.utc).replace(tzinfo=None),
    }


def serialize_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")
    if isinstance(value, date):
        return value.isoformat()
    return value


def serialize_row(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if row is None:
        return None
    return {key: serialize_value(value) for key, value in row.items()}


class Sistem:
    def __init__(self, model: Model, config: Settings = settings) -> None:
        self.model = model
        self.config = config

    def process_payload(self, payload: bytes | str | dict[str, Any]) -> int:
        return self.model.insert_reading(validate_payload(payload))

    def data_status(self, latest: dict[str, Any] | None) -> dict[str, Any]:
        if latest is None:
            return {
                "state": "no_data",
                "age_seconds": None,
                "stale_after_seconds": self.config.stale_after_seconds,
            }
        received_at = latest["received_at"].replace(tzinfo=timezone.utc)
        age = max(0, int((datetime.now(timezone.utc) - received_at).total_seconds()))
        return {
            "state": "stale" if age > self.config.stale_after_seconds else "online",
            "age_seconds": age,
            "stale_after_seconds": self.config.stale_after_seconds,
        }

    def latest(self) -> dict[str, Any]:
        latest = self.model.latest_reading()
        return {
            "latest_reading": serialize_row(latest),
            "data_status": self.data_status(latest),
        }

    def history(
        self, start: datetime | None = None, end: datetime | None = None, limit: int = 120
    ) -> dict[str, Any]:
        readings = [serialize_row(row) for row in self.model.readings(start=start, end=end, limit=limit)]
        return {"readings": readings, "count": len(readings)}

    def dashboard_summary(self) -> dict[str, Any]:
        start, end = self.local_date_range(datetime.now(self.config.timezone).date())
        latest = self.model.latest_reading()
        recent = self.model.readings(limit=8)
        statistics = self.clean_statistics(self.model.statistics(start, end))
        notifications = self.model.notifications(8)
        return {
            "latest_reading": serialize_row(latest),
            "data_status": self.data_status(latest),
            "today_statistics": statistics,
            "abnormal_counts": {
                "temperature": statistics["temperature_abnormal_count"],
                "gas": statistics["gas_abnormal_count"],
                "dht_problem": statistics["dht_problem_count"],
            },
            "notifications": [serialize_row(row) for row in notifications],
            "recent_readings": [serialize_row(row) for row in recent],
        }

    def report_summary(self, start: datetime, end: datetime) -> dict[str, Any]:
        stats = self.clean_statistics(self.model.statistics(start, end))
        return {
            "range": {"start": serialize_value(start), "end": serialize_value(end)},
            "statistics": {
                "temperature": {
                    "minimum": stats["temperature_min"],
                    "maximum": stats["temperature_max"],
                    "average": stats["temperature_avg"],
                },
                "humidity": {
                    "minimum": stats["humidity_min"],
                    "maximum": stats["humidity_max"],
                    "average": stats["humidity_avg"],
                },
                "gas": {
                    "minimum": stats["gas_min"],
                    "maximum": stats["gas_max"],
                    "average": stats["gas_avg"],
                },
            },
            "counts": {
                "total_readings": stats["total_readings"],
                "temperature_abnormal": stats["temperature_abnormal_count"],
                "gas_abnormal": stats["gas_abnormal_count"],
                "dht_problem": stats["dht_problem_count"],
                "buzzer_active": stats["buzzer_active_count"],
                "buzzer_inconsistent": stats["buzzer_inconsistent_count"],
            },
        }

    def report_readings(
        self, start: datetime, end: datetime, page: int, page_size: int
    ) -> dict[str, Any]:
        total = self.model.count_readings(start, end)
        rows = self.model.readings(
            start=start,
            end=end,
            limit=page_size,
            offset=(page - 1) * page_size,
        )
        total_pages = math.ceil(total / page_size) if total else 0
        return {
            "readings": [serialize_row(row) for row in rows],
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total_items": total,
                "total_pages": total_pages,
            },
        }

    def parse_report_range(self, start_date: str | None, end_date: str | None) -> tuple[datetime, datetime]:
        try:
            local_today = datetime.now(self.config.timezone).date()
            start_value = date.fromisoformat(start_date) if start_date else local_today
            end_value = date.fromisoformat(end_date) if end_date else local_today
        except ValueError as exc:
            raise PayloadError("Format tanggal harus YYYY-MM-DD.") from exc
        if end_value < start_value:
            raise PayloadError("Tanggal akhir tidak boleh sebelum tanggal mulai.")
        if (end_value - start_value).days >= 7:
            raise PayloadError("Rentang laporan maksimal tujuh hari.")
        return self.local_date_range(start_value, end_value)

    def local_date_range(
        self, start_date: date, end_date: date | None = None
    ) -> tuple[datetime, datetime]:
        end_date = end_date or start_date
        local_start = datetime.combine(start_date, time.min, self.config.timezone)
        local_end = datetime.combine(end_date + timedelta(days=1), time.min, self.config.timezone)
        return local_start.astimezone(timezone.utc), local_end.astimezone(timezone.utc)

    @staticmethod
    def clean_statistics(stats: dict[str, Any]) -> dict[str, Any]:
        keys = (
            "total_readings",
            "temperature_min",
            "temperature_max",
            "temperature_avg",
            "humidity_min",
            "humidity_max",
            "humidity_avg",
            "gas_min",
            "gas_max",
            "gas_avg",
            "temperature_abnormal_count",
            "gas_abnormal_count",
            "dht_problem_count",
            "buzzer_active_count",
            "buzzer_inconsistent_count",
        )
        cleaned: dict[str, Any] = {}
        for key in keys:
            value = stats.get(key)
            if key.endswith("_count") or key == "total_readings":
                cleaned[key] = int(value or 0)
            else:
                cleaned[key] = None if value is None else round(float(value), 2)
        return cleaned

    def export_excel(self, start: datetime, end: datetime) -> BytesIO:
        summary = self.report_summary(start, end)
        rows = self.model.readings(start=start, end=end, limit=302_400, ascending=True)
        events = self.model.abnormal_events(start, end)

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Ringkasan"
        summary_sheet.append(["Laporan Sensor Smart Maggot Farming"])
        summary_sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        summary_sheet["A1"].fill = PatternFill("solid", fgColor="087343")
        summary_sheet.append(["Dibuat", datetime.now(self.config.timezone).isoformat(timespec="seconds")])
        summary_sheet.append(["Rentang UTC", f"{serialize_value(start)} - {serialize_value(end)}"])
        summary_sheet.append([])
        for sensor, values in summary["statistics"].items():
            summary_sheet.append([sensor.title(), values["minimum"], values["maximum"], values["average"]])
        summary_sheet.append([])
        for key, value in summary["counts"].items():
            summary_sheet.append([key.replace("_", " ").title(), value])

        reading_sheet = workbook.create_sheet("Data Sensor")
        headers = [
            "Waktu UTC",
            "Suhu (°C)",
            "Kelembapan (%)",
            "Gas",
            "Buzzer",
            "Suhu Abnormal",
            "Gas Abnormal",
            "DHT Bermasalah",
            "Buzzer Tidak Konsisten",
        ]
        reading_sheet.append(headers)
        for row in rows:
            reading_sheet.append(
                [
                    serialize_value(row["received_at"]),
                    float(row["temperature"]),
                    float(row["humidity"]),
                    row["gas"],
                    row["buzzer"],
                    bool(row["temperature_abnormal"]),
                    bool(row["gas_abnormal"]),
                    bool(row["has_problem"]),
                    bool(row["buzzer_inconsistent"]),
                ]
            )

        event_sheet = workbook.create_sheet("Kondisi Abnormal")
        event_sheet.append(["Waktu UTC", "Tipe", "Tingkat", "Pesan"])
        for event in events:
            event_sheet.append(
                [
                    serialize_value(event["created_at"]),
                    event["notification_type"],
                    event["severity"],
                    event["message"],
                ]
            )

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.font = Font(bold=True)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def export_pdf(self, start: datetime, end: datetime) -> BytesIO:
        summary = self.report_summary(start, end)
        rows = self.model.readings(start=start, end=end, limit=1000, ascending=True)
        output = BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            leftMargin=1.2 * cm,
            rightMargin=1.2 * cm,
            topMargin=1.2 * cm,
            bottomMargin=1.2 * cm,
        )
        styles = getSampleStyleSheet()
        story = [
            Paragraph("Laporan Sensor Smart Maggot Farming", styles["Title"]),
            Paragraph(
                f"Rentang UTC: {serialize_value(start)} sampai {serialize_value(end)}",
                styles["Normal"],
            ),
            Spacer(1, 0.3 * cm),
        ]
        stats_rows = [["Sensor", "Minimum", "Maksimum", "Rata-rata"]]
        for sensor, values in summary["statistics"].items():
            stats_rows.append(
                [sensor.title(), values["minimum"], values["maximum"], values["average"]]
            )
        stats_table = Table(stats_rows, colWidths=[5 * cm, 4 * cm, 4 * cm, 4 * cm])
        stats_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#087343")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbded2")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f7f3")]),
                ]
            )
        )
        story.extend([stats_table, Spacer(1, 0.5 * cm)])

        data_rows = [["Waktu UTC", "Suhu", "Lembap", "Gas", "Buzzer", "Status"]]
        for row in rows:
            status = []
            if row["has_problem"]:
                status.append("DHT bermasalah")
            if row["temperature_abnormal"]:
                status.append("Suhu abnormal")
            if row["gas_abnormal"]:
                status.append("Gas abnormal")
            if row["buzzer_inconsistent"]:
                status.append("Buzzer tidak konsisten")
            data_rows.append(
                [
                    serialize_value(row["received_at"]),
                    float(row["temperature"]),
                    float(row["humidity"]),
                    row["gas"],
                    row["buzzer"],
                    ", ".join(status) or "Normal",
                ]
            )
        data_table = Table(data_rows, repeatRows=1, colWidths=[5 * cm, 2.2 * cm, 2.2 * cm, 2 * cm, 2 * cm, 8 * cm])
        data_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#087343")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#dbe7df")),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5faf7")]),
                ]
            )
        )
        story.append(data_table)
        document.build(story)
        output.seek(0)
        return output
